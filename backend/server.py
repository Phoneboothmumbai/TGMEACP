from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, BackgroundTasks, Header
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import bcrypt
import jwt
from jwt.exceptions import InvalidTokenError
import aiosmtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import httpx
import aiofiles
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io
import openpyxl
from openpyxl import Workbook

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Secret
JWT_SECRET = os.environ.get('JWT_SECRET', 'applecare-activation-secret-key-2025')
JWT_ALGORITHM = "HS256"

# Upload directory
UPLOAD_DIR = ROOT_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)
INVOICE_DIR = ROOT_DIR / 'invoices'
INVOICE_DIR.mkdir(exist_ok=True)

app = FastAPI(title="AppleCare+ Activation System")
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class AppleCarePlan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str = ""
    part_code: str = ""
    sku: str = ""
    description: str = ""
    mrp: Optional[float] = None
    active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AppleCarePlanCreate(BaseModel):
    name: str = ""
    part_code: str = ""
    sku: str = ""
    description: str = ""
    mrp: Optional[float] = None

class ActivationRequestCreate(BaseModel):
    dealer_name: str
    dealer_mobile: str
    dealer_email: EmailStr
    customer_name: str
    customer_mobile: str
    customer_email: EmailStr
    model_id: str
    serial_number: str
    plan_id: str
    device_activation_date: str

class ActivationRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    dealer_name: str
    dealer_mobile: str
    dealer_email: str = ""
    customer_name: str
    customer_mobile: str
    customer_email: str
    model_id: str
    serial_number: str
    plan_id: str
    plan_name: Optional[str] = ""
    plan_part_code: Optional[str] = ""
    plan_sku: Optional[str] = ""
    plan_mrp: Optional[float] = None
    device_activation_date: str
    billing_location: str = "F9B4869273B7"  # Hardcoded as per requirement
    payment_type: str = "Insta"  # Hardcoded as per requirement
    invoice_path: Optional[str] = None
    status: str = "pending"
    tgme_ticket_id: Optional[str] = None  # Renamed from osticket_id
    email_sent: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SettingsModel(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "main_settings"
    apple_email: str = ""  # Now supports comma-separated emails
    smtp_host: str = "smtp.gmail.com"
    smtp_port: int = 587
    smtp_email: str = ""
    smtp_password: str = ""
    tgme_url: str = ""  # Renamed from osticket_url
    tgme_api_key: str = ""  # Renamed from osticket_api_key
    # Keep old field names for backward compatibility
    osticket_url: str = ""
    osticket_api_key: str = ""
    partner_name: str = ""
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SettingsUpdate(BaseModel):
    apple_email: Optional[str] = None
    smtp_host: Optional[str] = None
    smtp_port: Optional[int] = None
    smtp_email: Optional[str] = None
    smtp_password: Optional[str] = None
    tgme_url: Optional[str] = None
    tgme_api_key: Optional[str] = None
    osticket_url: Optional[str] = None
    osticket_api_key: Optional[str] = None
    partner_name: Optional[str] = None

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc).timestamp() + 86400 * 7
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(authorization: str = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Authorization header missing")
    
    try:
        token = authorization.replace("Bearer ", "")
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(data: UserCreate):
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    user = {
        "id": user_id,
        "email": data.email,
        "name": data.name,
        "password": hash_password(data.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    token = create_token(user_id, data.email)
    
    return TokenResponse(
        access_token=token,
        user=UserResponse(id=user_id, email=data.email, name=data.name)
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: UserLogin):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user or not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"], user["email"])
    return TokenResponse(
        access_token=token,
        user=UserResponse(id=user["id"], email=user["email"], name=user["name"])
    )

@api_router.post("/auth/change-password")
async def change_password(data: PasswordChange, user: dict = Depends(get_current_user)):
    user_doc = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    
    if not verify_password(data.current_password, user_doc["password"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password": hash_password(data.new_password)}}
    )
    return {"message": "Password changed successfully"}

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(id=user["id"], email=user["email"], name=user["name"])

# ==================== PLANS ROUTES ====================

@api_router.get("/plans", response_model=List[AppleCarePlan])
async def get_plans(active_only: bool = True, public: bool = False):
    # Public endpoint for form dropdown - no auth required when public=True
    query = {"active": True} if active_only else {}
    plans = await db.plans.find(query, {"_id": 0}).to_list(1000)
    for plan in plans:
        if isinstance(plan.get('created_at'), str):
            plan['created_at'] = datetime.fromisoformat(plan['created_at'])
    return plans

@api_router.post("/plans", response_model=AppleCarePlan)
async def create_plan(data: AppleCarePlanCreate, user: dict = Depends(get_current_user)):
    plan = AppleCarePlan(**data.model_dump())
    doc = plan.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.plans.insert_one(doc)
    return plan

@api_router.put("/plans/{plan_id}", response_model=AppleCarePlan)
async def update_plan(plan_id: str, data: AppleCarePlanCreate, user: dict = Depends(get_current_user)):
    update_data = data.model_dump()
    result = await db.plans.update_one({"id": plan_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    plan = await db.plans.find_one({"id": plan_id}, {"_id": 0})
    if isinstance(plan.get('created_at'), str):
        plan['created_at'] = datetime.fromisoformat(plan['created_at'])
    return plan

@api_router.delete("/plans/{plan_id}")
async def delete_plan(plan_id: str, user: dict = Depends(get_current_user)):
    result = await db.plans.update_one({"id": plan_id}, {"$set": {"active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    return {"message": "Plan deactivated"}

@api_router.get("/plans/sample")
async def download_sample_excel(user: dict = Depends(get_current_user)):
    """Download a sample Excel file for AppleCare+ plans upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "AppleCare+ Plans"
    
    # Headers
    headers = ["SKU", "Description", "MRP", "Part Code", "Plan Name"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Sample data
    sample_data = [
        ["S9732ZM/A", "AppleCare+ for iPhone 15", 14900, "SR182HN/A", "AppleCare+ for iPhone 15"],
        ["S9733ZM/A", "AppleCare+ for iPhone 15 Pro", 23900, "SR183HN/A", "AppleCare+ for iPhone 15 Pro"],
        ["S9734ZM/A", "AppleCare+ for iPhone 15 Pro Max", 23900, "SR184HN/A", "AppleCare+ for iPhone 15 Pro Max"],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    
    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=applecare_plans_sample.xlsx"}
    )

@api_router.post("/plans/upload")
async def upload_plans_excel(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload AppleCare+ plans from Excel file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1] if cell.value]
        header_map = {h.lower().strip(): idx for idx, h in enumerate(headers)}
        
        imported_count = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # Skip empty rows
                continue
                
            try:
                # Extract data based on header positions
                sku = str(row[header_map.get('sku', 0)] or '').strip()
                description = str(row[header_map.get('description', 1)] or '').strip()
                mrp_val = row[header_map.get('mrp', 2)]
                part_code = str(row[header_map.get('part code', 3)] or row[header_map.get('partcode', 3)] or '').strip()
                plan_name = str(row[header_map.get('plan name', 4)] or row[header_map.get('name', 4)] or '').strip()
                
                # Parse MRP
                mrp = None
                if mrp_val:
                    try:
                        mrp = float(str(mrp_val).replace(',', '').replace('₹', '').strip())
                    except ValueError:
                        pass
                
                # Skip if no SKU and no part code
                if not sku and not part_code:
                    continue
                
                # Check if plan with same SKU already exists
                existing = await db.plans.find_one({"$or": [{"sku": sku}, {"part_code": part_code}]})
                
                if existing:
                    # Update existing plan
                    await db.plans.update_one(
                        {"id": existing["id"]},
                        {"$set": {
                            "sku": sku or existing.get("sku", ""),
                            "description": description or existing.get("description", ""),
                            "mrp": mrp if mrp else existing.get("mrp"),
                            "part_code": part_code or existing.get("part_code", ""),
                            "name": plan_name or existing.get("name", ""),
                            "active": True
                        }}
                    )
                else:
                    # Create new plan
                    plan = AppleCarePlan(
                        sku=sku,
                        description=description,
                        mrp=mrp,
                        part_code=part_code,
                        name=plan_name
                    )
                    doc = plan.model_dump()
                    doc['created_at'] = doc['created_at'].isoformat()
                    await db.plans.insert_one(doc)
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        return {
            "message": f"Successfully imported {imported_count} plans",
            "imported_count": imported_count,
            "errors": errors[:10] if errors else []  # Return first 10 errors
        }
        
    except Exception as e:
        logger.error(f"Excel upload error: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to process Excel file: {str(e)}")

# ==================== SETTINGS ROUTES ====================

@api_router.get("/settings", response_model=SettingsModel)
async def get_settings(user: dict = Depends(get_current_user)):
    settings = await db.settings.find_one({"id": "main_settings"}, {"_id": 0})
    if not settings:
        default_settings = SettingsModel()
        doc = default_settings.model_dump()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.settings.insert_one(doc)
        return default_settings
    if isinstance(settings.get('updated_at'), str):
        settings['updated_at'] = datetime.fromisoformat(settings['updated_at'])
    return settings

@api_router.put("/settings", response_model=SettingsModel)
async def update_settings(data: SettingsUpdate, user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.settings.update_one(
        {"id": "main_settings"},
        {"$set": update_data},
        upsert=True
    )
    settings = await db.settings.find_one({"id": "main_settings"}, {"_id": 0})
    if isinstance(settings.get('updated_at'), str):
        settings['updated_at'] = datetime.fromisoformat(settings['updated_at'])
    return settings

# ==================== PDF GENERATION ====================

import random
import string

# Indian mobile shop names for realistic invoices
SHOP_NAMES = [
    "TechZone Mobile Hub",
    "Digital Dreams Electronics",
    "Mobile Planet India",
    "SmartCell Solutions",
    "iWorld Mobile Store",
    "Galaxy Tech Mart",
    "Prime Mobile House",
    "Supreme Electronics",
    "NextGen Mobile Shop",
    "City Mobile Center",
    "Metro Tech Store",
    "Royal Mobile Emporium",
    "Star Mobile Point",
    "Express Mobile Mart",
    "Horizon Electronics"
]

# Indian addresses for realistic invoices
SHOP_ADDRESSES = [
    {"address": "Shop 12, Linking Road, Bandra West", "city": "Mumbai", "pin": "400050", "state": "27-Maharashtra", "phone": "9820098200"},
    {"address": "45, MG Road, Camp Area", "city": "Pune", "pin": "411001", "state": "27-Maharashtra", "phone": "9890123456"},
    {"address": "23, Brigade Road, Near Commercial Street", "city": "Bangalore", "pin": "560001", "state": "29-Karnataka", "phone": "9845012345"},
    {"address": "F-15, Connaught Place, Block F", "city": "New Delhi", "pin": "110001", "state": "07-Delhi", "phone": "9811234567"},
    {"address": "Shop 8, Anna Salai, Opposite Express Avenue", "city": "Chennai", "pin": "600002", "state": "33-Tamil Nadu", "phone": "9841234567"},
    {"address": "102, CG Road, Navrangpura", "city": "Ahmedabad", "pin": "380009", "state": "24-Gujarat", "phone": "9825678901"},
    {"address": "28, Park Street, Near Park Circus", "city": "Kolkata", "pin": "700017", "state": "19-West Bengal", "phone": "9830123456"},
    {"address": "5, Sector 17, Plaza Market", "city": "Chandigarh", "pin": "160017", "state": "04-Chandigarh", "phone": "9815012345"},
    {"address": "Shop 33, Hazratganj, Main Road", "city": "Lucknow", "pin": "226001", "state": "09-Uttar Pradesh", "phone": "9839012345"},
    {"address": "201, MI Road, Near Statue Circle", "city": "Jaipur", "pin": "302001", "state": "08-Rajasthan", "phone": "9829012345"}
]

# Product pricing based on AppleCare+ plan description
PRODUCT_PRICING = {
    "macbook air": {"name": "MacBook Air", "price": 80000},
    "macbook pro": {"name": "MacBook Pro", "price": 169900},
    "iphone": {"name": "iPhone", "price": 79900},
    "iphone pro": {"name": "iPhone Pro", "price": 134900},
    "iphone pro max": {"name": "iPhone Pro Max", "price": 149900},
    "ipad": {"name": "iPad", "price": 39900},
    "ipad air": {"name": "iPad Air", "price": 59900},
    "ipad pro": {"name": "iPad Pro", "price": 89900},
    "imac": {"name": "iMac", "price": 134900},
    "apple watch": {"name": "Apple Watch", "price": 44900},
    "apple watch ultra": {"name": "Apple Watch Ultra", "price": 89900},
    "airpods": {"name": "AirPods", "price": 14900},
    "airpods pro": {"name": "AirPods Pro", "price": 24900}
}

def detect_product_from_plan(plan_name: str, plan_description: str) -> dict:
    """Detect the Apple product from the AppleCare+ plan name/description"""
    combined = (plan_name + " " + plan_description).lower()
    
    # Check for specific products (order matters - more specific first)
    if "pro max" in combined and "iphone" in combined:
        return PRODUCT_PRICING["iphone pro max"]
    elif "iphone" in combined and "pro" in combined:
        return PRODUCT_PRICING["iphone pro"]
    elif "iphone" in combined:
        return PRODUCT_PRICING["iphone"]
    elif "macbook pro" in combined or "mac pro" in combined:
        return PRODUCT_PRICING["macbook pro"]
    elif "macbook air" in combined or "mac air" in combined:
        return PRODUCT_PRICING["macbook air"]
    elif "macbook" in combined or "mac" in combined:
        return PRODUCT_PRICING["macbook air"]  # Default Mac
    elif "ipad pro" in combined:
        return PRODUCT_PRICING["ipad pro"]
    elif "ipad air" in combined:
        return PRODUCT_PRICING["ipad air"]
    elif "ipad" in combined:
        return PRODUCT_PRICING["ipad"]
    elif "imac" in combined:
        return PRODUCT_PRICING["imac"]
    elif "watch ultra" in combined:
        return PRODUCT_PRICING["apple watch ultra"]
    elif "watch" in combined:
        return PRODUCT_PRICING["apple watch"]
    elif "airpods pro" in combined:
        return PRODUCT_PRICING["airpods pro"]
    elif "airpods" in combined:
        return PRODUCT_PRICING["airpods"]
    else:
        # Default to iPhone if can't detect
        return PRODUCT_PRICING["iphone"]

def num_to_words_indian(num: int) -> str:
    """Convert number to Indian currency words"""
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
            "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    
    if num == 0:
        return "Zero"
    
    def two_digits(n):
        if n < 20:
            return ones[n]
        return tens[n // 10] + (" " + ones[n % 10] if n % 10 else "")
    
    def three_digits(n):
        if n < 100:
            return two_digits(n)
        return ones[n // 100] + " Hundred" + (" " + two_digits(n % 100) if n % 100 else "")
    
    result = ""
    if num >= 10000000:  # Crore
        result += three_digits(num // 10000000) + " Crore "
        num %= 10000000
    if num >= 100000:  # Lakh
        result += two_digits(num // 100000) + " Lakh "
        num %= 100000
    if num >= 1000:  # Thousand
        result += two_digits(num // 1000) + " Thousand "
        num %= 1000
    if num >= 100:  # Hundred
        result += ones[num // 100] + " Hundred "
        num %= 100
    if num > 0:
        result += two_digits(num)
    
    return result.strip()

def format_indian_currency(amount: float) -> str:
    """Format amount in Indian currency format (e.g., 1,70,800.00)"""
    amount_str = f"{amount:,.2f}"
    parts = amount_str.split(".")
    integer_part = parts[0].replace(",", "")
    decimal_part = parts[1] if len(parts) > 1 else "00"
    
    # Indian numbering system
    if len(integer_part) <= 3:
        formatted = integer_part
    else:
        formatted = integer_part[-3:]
        integer_part = integer_part[:-3]
        while integer_part:
            formatted = integer_part[-2:] + "," + formatted
            integer_part = integer_part[:-2]
    
    return f"₹ {formatted}.{decimal_part}"

async def generate_invoice_pdf(request_data: dict, filename: str) -> str:
    filepath = INVOICE_DIR / filename
    
    # Random shop details
    shop_name = random.choice(SHOP_NAMES)
    shop_addr = random.choice(SHOP_ADDRESSES)
    invoice_number = ''.join(random.choices(string.digits, k=4))
    
    # Get activation date for invoice date
    invoice_date = request_data.get('device_activation_date', datetime.now().strftime("%d-%m-%Y"))
    if "-" in invoice_date and len(invoice_date.split("-")[0]) == 4:
        # Convert from YYYY-MM-DD to DD-MM-YYYY
        parts = invoice_date.split("-")
        invoice_date = f"{parts[2]}-{parts[1]}-{parts[0]}"
    
    # Detect product from AppleCare+ plan
    plan_name = request_data.get('plan_name', '')
    plan_description = request_data.get('plan_description', plan_name)
    product_info = detect_product_from_plan(plan_name, plan_description)
    
    # Get AppleCare+ price (MRP)
    applecare_price = request_data.get('plan_mrp', 0) or 14900  # Default AppleCare+ price
    product_price = product_info["price"]
    
    # Calculate tax (18% GST inclusive)
    # For inclusive GST: Base = Total / 1.18, GST = Total - Base
    product_base = round(product_price / 1.18, 2)
    product_gst = round(product_price - product_base, 2)
    
    applecare_base = round(applecare_price / 1.18, 2)
    applecare_gst = round(applecare_price - applecare_base, 2)
    
    total_amount = product_price + applecare_price
    total_base = product_base + applecare_base
    total_gst = product_gst + applecare_gst
    cgst = round(total_gst / 2, 2)
    sgst = round(total_gst / 2, 2)
    
    # Create PDF
    doc = SimpleDocTemplate(str(filepath), pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1a1a1a'), spaceAfter=5)
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#666666'))
    bold_style = ParagraphStyle('Bold', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold')
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9)
    small_style = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#666666'))
    
    # Header Section - Company Info and Sale Order
    header_data = [
        [
            [
                Paragraph(f"<b>{shop_name}</b>", title_style),
                Paragraph(f"{shop_addr['address']}<br/>{shop_addr['city']}, {shop_addr['pin']}<br/>Phone: {shop_addr['phone']}<br/>State: {shop_addr['state']}", header_style)
            ],
            [
                Paragraph("<b>Sale Order</b>", ParagraphStyle('SO', parent=styles['Heading2'], fontSize=14, alignment=2)),
                Paragraph(f"<b>Invoice No:</b> {invoice_number}<br/><b>Date:</b> {invoice_date}", ParagraphStyle('SODetails', parent=styles['Normal'], fontSize=9, alignment=2))
            ]
        ]
    ]
    
    header_table = Table(header_data, colWidths=[4*inch, 3*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Bill To Section
    elements.append(Paragraph("<b>Bill To</b>", bold_style))
    customer_info = f"{request_data.get('customer_name', '')}<br/>"
    customer_info += f"Phone: {request_data.get('customer_mobile', '')}<br/>"
    customer_info += f"Email: {request_data.get('customer_email', '')}"
    elements.append(Paragraph(customer_info, normal_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Product Table
    product_table_data = [
        ["Item name", "HSN/SAC", "Qty", "Price/Unit", "GST", "Amount"]
    ]
    
    # Product row
    serial_no = request_data.get('serial_number', '')
    product_name_with_serial = f"<b>{product_info['name'].upper()}</b><br/><font size=7>Serial No.: {serial_no}</font>"
    product_table_data.append([
        Paragraph(product_name_with_serial, normal_style),
        "85171290",  # HSN code for mobile phones
        "1",
        format_indian_currency(product_base),
        f"{format_indian_currency(product_gst)}\n(18%)",
        format_indian_currency(product_price)
    ])
    
    # AppleCare+ row - also include serial number
    applecare_name = plan_name if plan_name else f"AppleCare+ for {product_info['name']}"
    applecare_name_with_serial = f"<b>{applecare_name}</b><br/><font size=7>Serial No.: {serial_no}</font>"
    product_table_data.append([
        Paragraph(applecare_name_with_serial, normal_style),
        "998716",  # SAC code for warranty services
        "1",
        format_indian_currency(applecare_base),
        f"{format_indian_currency(applecare_gst)}\n(18%)",
        format_indian_currency(applecare_price)
    ])
    
    # Total row
    product_table_data.append([
        "", "", "", "Total",
        format_indian_currency(total_gst),
        format_indian_currency(total_amount)
    ])
    
    product_table = Table(product_table_data, colWidths=[2.5*inch, 0.8*inch, 0.5*inch, 1.2*inch, 1*inch, 1.2*inch])
    product_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f5f5f5')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('FONTNAME', (3, -1), (3, -1), 'Helvetica-Bold'),
    ]))
    elements.append(product_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Amount in words
    amount_words = num_to_words_indian(int(total_amount))
    elements.append(Paragraph(f"<b>Invoice Amount in Words:</b> {amount_words} Rupees only", normal_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Amount Summary
    amount_summary_data = [
        ["Total:", format_indian_currency(total_amount)],
        ["Received:", format_indian_currency(0)],
        ["Balance:", format_indian_currency(total_amount)]
    ]
    
    amount_summary = Table(amount_summary_data, colWidths=[1*inch, 1.5*inch])
    amount_summary.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(amount_summary)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tax Breakdown
    elements.append(Paragraph("<b>Tax Breakdown</b>", bold_style))
    tax_data = [
        ["HSN/SAC", "Taxable Amt", "CGST Rate", "CGST Amt", "SGST Rate", "SGST Amt", "Total Tax"]
    ]
    
    # Product tax
    tax_data.append([
        "85171290",
        format_indian_currency(product_base),
        "9%",
        format_indian_currency(product_gst / 2),
        "9%",
        format_indian_currency(product_gst / 2),
        format_indian_currency(product_gst)
    ])
    
    # AppleCare+ tax
    tax_data.append([
        "998716",
        format_indian_currency(applecare_base),
        "9%",
        format_indian_currency(applecare_gst / 2),
        "9%",
        format_indian_currency(applecare_gst / 2),
        format_indian_currency(applecare_gst)
    ])
    
    # Tax totals
    tax_data.append([
        "Total",
        format_indian_currency(total_base),
        "",
        format_indian_currency(cgst),
        "",
        format_indian_currency(sgst),
        format_indian_currency(total_gst)
    ])
    
    tax_table = Table(tax_data, colWidths=[0.8*inch, 1.1*inch, 0.7*inch, 0.9*inch, 0.7*inch, 0.9*inch, 0.9*inch])
    tax_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f5f5f5')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(tax_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Footer
    elements.append(Paragraph("<b>Terms and conditions</b>", bold_style))
    elements.append(Paragraph("Thanks for doing business with us!", small_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Authorized Signatory
    footer_data = [
        ["", f"For {shop_name}"],
        ["", ""],
        ["", ""],
        ["", "Authorized Signatory"]
    ]
    footer_table = Table(footer_data, colWidths=[4.5*inch, 2.5*inch])
    footer_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (1, -1), (1, -1), 'Helvetica'),
    ]))
    elements.append(footer_table)
    
    doc.build(elements)
    return str(filepath)

# ==================== EMAIL SERVICE ====================

async def send_activation_email(request_data: dict, invoice_path: Optional[str] = None, ticket_id: Optional[str] = None):
    settings = await db.settings.find_one({"id": "main_settings"}, {"_id": 0})
    if not settings or not settings.get('smtp_email') or not settings.get('apple_email'):
        logger.warning("Email settings not configured")
        return False
    
    # Parse multiple email addresses (comma-separated)
    apple_emails = [e.strip() for e in settings['apple_email'].split(',') if e.strip()]
    if not apple_emails:
        logger.warning("No valid Apple email addresses configured")
        return False
    
    msg = MIMEMultipart()
    msg['From'] = settings['smtp_email']
    msg['To'] = ', '.join(apple_emails)  # Join multiple recipients
    
    # Email subject format: AppleCare+ for (Customer name) #(OSTICKETID)
    customer_name = request_data.get('customer_name', 'Customer')
    if ticket_id:
        msg['Subject'] = f"AppleCare+ for {customer_name} #{ticket_id}"
    else:
        msg['Subject'] = f"AppleCare+ for {customer_name}"
    
    # Build email body with tabular format
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
    <h2>AppleCare+ Activation Request</h2>
    <p>Please find the activation details below:</p>
    <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse;">
        <tr style="background-color: #f2f2f2;">
            <th>IMEI/Serial</th>
            <th>NAME</th>
            <th>EMAIL ID</th>
            <th>MOBILE NO</th>
            <th>Plan Part Code</th>
            <th>Device DOP</th>
            <th>Billing Location</th>
            <th>Payment Type</th>
            <th>Plan Name</th>
            <th>Partner Name</th>
        </tr>
        <tr>
            <td>{request_data.get('serial_number', '')}</td>
            <td>{request_data.get('customer_name', '')}</td>
            <td>{request_data.get('customer_email', '')}</td>
            <td>{request_data.get('customer_mobile', '')}</td>
            <td>{request_data.get('plan_sku', '') or request_data.get('plan_part_code', '')}</td>
            <td>{request_data.get('device_activation_date', '')}</td>
            <td>{request_data.get('billing_location', 'F9B4869273B7')}</td>
            <td>{request_data.get('payment_type', 'Insta')}</td>
            <td>AppleCare+</td>
            <td>{settings.get('partner_name', '')}</td>
        </tr>
    </table>
    <br>
    <p>Best regards,<br>{settings.get('partner_name', 'Partner')}</p>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html_body, 'html'))
    
    # Attach invoice if exists
    if invoice_path and os.path.exists(invoice_path):
        with open(invoice_path, 'rb') as f:
            part = MIMEBase('application', 'pdf')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment; filename=invoice.pdf')
            msg.attach(part)
    
    try:
        await aiosmtplib.send(
            msg,
            recipients=apple_emails,  # Explicitly pass all recipients
            hostname=settings.get('smtp_host', 'smtp.gmail.com'),
            port=settings.get('smtp_port', 587),
            username=settings['smtp_email'],
            password=settings['smtp_password'],
            start_tls=True
        )
        logger.info(f"Email sent successfully to {', '.join(apple_emails)}")
        return True
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        return False

# ==================== TGME SUPPORT TICKET SERVICE ====================

async def create_tgme_ticket(request_data: dict):
    """Create a TGME Support Ticket (formerly osTicket) using DEALER details"""
    settings = await db.settings.find_one({"id": "main_settings"}, {"_id": 0})
    
    # Support both new and old field names for backward compatibility
    tgme_url = settings.get('tgme_url') or settings.get('osticket_url') if settings else None
    tgme_api_key = settings.get('tgme_api_key') or settings.get('osticket_api_key') if settings else None
    
    if not tgme_url or not tgme_api_key:
        logger.warning("TGME Support Ticket settings not configured")
        return None
    
    # Build comprehensive ticket body with ALL form details
    ticket_body = f"""
================================================================================
                        AppleCare+ ACTIVATION REQUEST
================================================================================

DEALER INFORMATION (Ticket Raised By)
--------------------------------------------------------------------------------
Dealer Name:          {request_data.get('dealer_name', '')}
Dealer Email:         {request_data.get('dealer_email', '')}
Dealer Mobile:        {request_data.get('dealer_mobile', '')}

CUSTOMER INFORMATION
--------------------------------------------------------------------------------
Customer Name:        {request_data.get('customer_name', '')}
Customer Email:       {request_data.get('customer_email', '')}
Customer Mobile:      {request_data.get('customer_mobile', '')}

DEVICE INFORMATION
--------------------------------------------------------------------------------
Model ID:             {request_data.get('model_id', '')}
Serial Number/IMEI:   {request_data.get('serial_number', '')}
Activation Date:      {request_data.get('device_activation_date', '')}

APPLECARE+ PLAN DETAILS
--------------------------------------------------------------------------------
Plan Name:            {request_data.get('plan_name', '')}
Plan SKU:             {request_data.get('plan_sku', '')}
Plan Part Code:       {request_data.get('plan_part_code', '')}
Plan MRP:             ₹{request_data.get('plan_mrp', 'N/A')}

INTERNAL REFERENCE
--------------------------------------------------------------------------------
Billing Location:     {request_data.get('billing_location', 'F9B4869273B7')}
Payment Type:         {request_data.get('payment_type', 'Insta')}
Request ID:           {request_data.get('id', '')}

================================================================================
"""
    
    # Use DEALER details for ticket creation (not customer)
    ticket_data = {
        "name": request_data.get('dealer_name', ''),
        "email": request_data.get('dealer_email', ''),
        "phone": request_data.get('dealer_mobile', ''),
        "subject": f"AppleCare+ Activation - {request_data.get('customer_name', '')} - {request_data.get('serial_number', '')}",
        "message": ticket_body,
        "topicId": "1"
    }
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                f"{tgme_url}/api/tickets.json",
                json=ticket_data,
                headers={
                    "X-API-Key": tgme_api_key,
                    "Content-Type": "application/json"
                }
            )
            if response.status_code == 201:
                ticket_id = response.text.strip('"')
                logger.info(f"TGME Support Ticket created: {ticket_id}")
                return ticket_id
            else:
                logger.error(f"TGME Support Ticket creation failed: {response.text}")
                return None
    except Exception as e:
        logger.error(f"TGME Support Ticket error: {e}")
        return None

# ==================== ACTIVATION REQUESTS ROUTES ====================

@api_router.get("/activation-requests", response_model=List[ActivationRequest])
async def get_activation_requests(status: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = {}
    if status:
        query["status"] = status
    requests = await db.activation_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for req in requests:
        if isinstance(req.get('created_at'), str):
            req['created_at'] = datetime.fromisoformat(req['created_at'])
        if isinstance(req.get('updated_at'), str):
            req['updated_at'] = datetime.fromisoformat(req['updated_at'])
    return requests

@api_router.get("/activation-requests/{request_id}", response_model=ActivationRequest)
async def get_activation_request(request_id: str, user: dict = Depends(get_current_user)):
    req = await db.activation_requests.find_one({"id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if isinstance(req.get('created_at'), str):
        req['created_at'] = datetime.fromisoformat(req['created_at'])
    if isinstance(req.get('updated_at'), str):
        req['updated_at'] = datetime.fromisoformat(req['updated_at'])
    return req

@api_router.post("/activation-requests", response_model=ActivationRequest)
async def create_activation_request(
    data: ActivationRequestCreate,
    background_tasks: BackgroundTasks
):
    # Public endpoint - no auth required
    # Get plan details
    plan = await db.plans.find_one({"id": data.plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=400, detail="Invalid plan selected")
    
    # Get plan details - prefer description over name, and sku over part_code
    # (Excel upload sometimes puts wrong values in name/part_code fields)
    plan_name_value = plan.get('description', '') or plan.get('name', '')
    plan_sku_value = plan.get('sku', '') or ''
    # Use SKU as part_code if part_code looks like a number (bad data)
    plan_part_code_value = plan.get('part_code', '')
    try:
        float(plan_part_code_value)  # If it's a number, it's bad data
        plan_part_code_value = plan_sku_value  # Use SKU instead
    except (ValueError, TypeError):
        pass  # It's a valid string, keep it
    
    request_obj = ActivationRequest(
        **data.model_dump(),
        plan_name=plan_name_value,
        plan_part_code=plan_part_code_value,
        plan_sku=plan_sku_value,
        plan_mrp=plan.get('mrp'),
        billing_location="F9B4869273B7",  # Hardcoded
        payment_type="Insta"  # Hardcoded
    )
    
    doc = request_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    # Generate invoice PDF
    invoice_filename = f"invoice_{request_obj.id}.pdf"
    invoice_path = await generate_invoice_pdf(doc, invoice_filename)
    doc['invoice_path'] = invoice_path
    
    await db.activation_requests.insert_one(doc)
    
    # Background tasks: create TGME ticket and send email
    background_tasks.add_task(process_activation_request, request_obj.id)
    
    return request_obj

async def process_activation_request(request_id: str):
    """Background task to create TGME Support Ticket and send email"""
    req = await db.activation_requests.find_one({"id": request_id}, {"_id": 0})
    if not req:
        return
    
    # Create TGME Support Ticket FIRST to get the ticket ID
    ticket_id = await create_tgme_ticket(req)
    if ticket_id:
        await db.activation_requests.update_one(
            {"id": request_id},
            {"$set": {"tgme_ticket_id": ticket_id, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    # Send email to Apple with ticket ID in subject
    email_sent = await send_activation_email(req, req.get('invoice_path'), ticket_id)
    await db.activation_requests.update_one(
        {"id": request_id},
        {"$set": {"email_sent": email_sent, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

@api_router.put("/activation-requests/{request_id}/status")
async def update_request_status(request_id: str, status: str, user: dict = Depends(get_current_user)):
    valid_statuses = ["pending", "email_sent", "payment_pending", "activated", "cancelled"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    result = await db.activation_requests.update_one(
        {"id": request_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    return {"message": "Status updated"}

@api_router.post("/activation-requests/{request_id}/resend-email")
async def resend_email(request_id: str, background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    req = await db.activation_requests.find_one({"id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Include ticket ID in email subject when resending
    ticket_id = req.get('tgme_ticket_id') or req.get('osticket_id')
    background_tasks.add_task(send_activation_email, req, req.get('invoice_path'), ticket_id)
    return {"message": "Email resend queued"}

@api_router.get("/activation-requests/{request_id}/invoice")
async def download_invoice(request_id: str, authorization: str = None):
    # Accept token from query parameter for file downloads (browsers can't send headers for direct links)
    if not authorization:
        raise HTTPException(status_code=401, detail="Authorization required")
    
    try:
        token = authorization.replace("Bearer ", "")
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
    except InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    req = await db.activation_requests.find_one({"id": request_id}, {"_id": 0})
    if not req or not req.get('invoice_path'):
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if not os.path.exists(req['invoice_path']):
        raise HTTPException(status_code=404, detail="Invoice file not found")
    
    return FileResponse(
        req['invoice_path'],
        media_type='application/pdf',
        filename=f"invoice_{request_id}.pdf"
    )

# ==================== FILE UPLOAD ====================

@api_router.post("/upload-invoice/{request_id}")
async def upload_invoice(request_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files allowed")
    
    filename = f"uploaded_invoice_{request_id}.pdf"
    filepath = UPLOAD_DIR / filename
    
    async with aiofiles.open(filepath, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    await db.activation_requests.update_one(
        {"id": request_id},
        {"$set": {"invoice_path": str(filepath), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Invoice uploaded", "path": str(filepath)}

# ==================== DASHBOARD STATS ====================

@api_router.get("/stats")
async def get_stats(user: dict = Depends(get_current_user)):
    total = await db.activation_requests.count_documents({})
    pending = await db.activation_requests.count_documents({"status": "pending"})
    activated = await db.activation_requests.count_documents({"status": "activated"})
    payment_pending = await db.activation_requests.count_documents({"status": "payment_pending"})
    
    return {
        "total": total,
        "pending": pending,
        "activated": activated,
        "payment_pending": payment_pending
    }

# ==================== HEALTH CHECK ====================

@api_router.get("/health")
async def health():
    return {"status": "healthy", "service": "applecare-activation"}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup():
    # Create default admin if not exists
    admin = await db.users.find_one({"email": "ck@motta.in"})
    if not admin:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": "ck@motta.in",
            "name": "Admin",
            "password": hash_password("Charu@123@"),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info("Admin user created")
    
    # Create some default plans if none exist
    plans_count = await db.plans.count_documents({})
    if plans_count == 0:
        default_plans = [
            {"id": str(uuid.uuid4()), "name": "AppleCare+", "part_code": "SR182HN/A", "description": "Standard AppleCare+ Protection", "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "AppleCare+ with Theft and Loss", "part_code": "SR183HN/A", "description": "AppleCare+ with Theft and Loss coverage", "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "AppleCare+ for Mac", "part_code": "SR184HN/A", "description": "AppleCare+ for Mac computers", "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "AppleCare+ for iPad", "part_code": "SR185HN/A", "description": "AppleCare+ for iPad devices", "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "AppleCare+ for Apple Watch", "part_code": "SR186HN/A", "description": "AppleCare+ for Apple Watch", "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.plans.insert_many(default_plans)
        logger.info("Default AppleCare+ plans created")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
